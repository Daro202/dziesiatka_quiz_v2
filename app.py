import os
import threading
from copy import deepcopy

import requests
from flask import Flask, jsonify, render_template, request, Response
from openpyxl import load_workbook


app = Flask(__name__)


DEFAULT_STATE = {
    "currentRound": "WARM UP",
    "question": "",
    "answer": "",
    "answerVisible": False,
    "currentPlayerId": None,
    "players": [],
    "allQuestions": [],
    "questionSet": "all",
    "usedQuestionIds": []
}

GAME_STATE = deepcopy(DEFAULT_STATE)
STATE_LOCK = threading.Lock()


def normalize_header(value):
    if value is None:
        return ""

    text = str(value).strip().lower()

    replacements = {
        "ą": "a",
        "ć": "c",
        "ę": "e",
        "ł": "l",
        "ń": "n",
        "ó": "o",
        "ś": "s",
        "ż": "z",
        "ź": "z",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    text = text.replace(" ", "_")
    text = text.replace("-", "_")

    return text


def get_value(row_data, *possible_names):
    for name in possible_names:
        normalized_name = normalize_header(name)
        if normalized_name in row_data:
            value = row_data.get(normalized_name)
            if value is None:
                return ""
            return str(value).strip()
    return ""


def get_flag(row_data, *possible_names):
    value = get_value(row_data, *possible_names)
    value = str(value).strip().lower()

    if value in ["1", "tak", "yes", "x", "true", "prawda"]:
        return "1"

    return ""


def load_questions_from_excel():
    file_path = os.path.join(app.root_path, "data", "quiz_questions.xlsx")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Nie znaleziono pliku: {file_path}")

    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [normalize_header(x) for x in rows[0]]

    questions = []

    for index, row in enumerate(rows[1:], start=1):
        row_data = {}

        for col_index, header in enumerate(headers):
            if not header:
                continue

            value = row[col_index] if col_index < len(row) else ""
            row_data[header] = value

        round_name = get_value(row_data, "round", "runda")
        question = get_value(row_data, "question", "pytanie")
        answer = get_value(row_data, "answer", "odpowiedz", "odpowiedź")

        if not question:
            continue

        question_id = get_value(row_data, "id", "question_id", "nr", "lp")
        if not question_id:
            question_id = str(index)

        questions.append({
            "id": str(question_id),
            "round": round_name,
            "question": question,
            "answer": answer,
            "lodz": get_flag(row_data, "lodz", "łódź"),
            "tychy": get_flag(row_data, "tychy"),
            "mosina": get_flag(row_data, "mosina"),
            "ostroleka": get_flag(row_data, "ostroleka", "ostrołęka"),
        })

    return questions


@app.route("/")
def home():
    return """
    <h1>Dziesiątka</h1>
    <p>Teleturniej wiedzy</p>
    <p><a href="admin">Panel prowadzącego</a></p>
    <p><a href="audience">Ekran publiczności</a></p>
    """


@app.route("/admin")
def admin():
    return render_template("admin.html")


@app.route("/audience")
def audience():
    return render_template("audience.html")


@app.route("/api/questions")
def api_questions():
    try:
        questions = load_questions_from_excel()
        return jsonify({
            "ok": True,
            "count": len(questions),
            "questions": questions
        })
    except Exception as error:
        return jsonify({
            "ok": False,
            "error": str(error)
        }), 500


@app.route("/api/state", methods=["GET"])
def api_get_state():
    with STATE_LOCK:
        return jsonify({
            "ok": True,
            "state": deepcopy(GAME_STATE)
        })


@app.route("/api/state", methods=["POST"])
def api_save_state():
    global GAME_STATE

    data = request.get_json(silent=True)

    if not isinstance(data, dict):
        return jsonify({
            "ok": False,
            "error": "Nieprawidłowy format stanu gry."
        }), 400

    with STATE_LOCK:
        new_state = deepcopy(DEFAULT_STATE)
        new_state.update(data)
        GAME_STATE = new_state

        return jsonify({
            "ok": True,
            "state": deepcopy(GAME_STATE)
        })


@app.route("/api/reset-state", methods=["POST"])
def api_reset_state():
    global GAME_STATE

    with STATE_LOCK:
        GAME_STATE = deepcopy(DEFAULT_STATE)

    return jsonify({
        "ok": True,
        "state": deepcopy(GAME_STATE)
    })


@app.route("/api/elevenlabs-tts", methods=["POST"])
def api_elevenlabs_tts():
    api_key = os.getenv("ELEVENLABS_API_KEY")

    if not api_key:
        return Response(
            "Brak zmiennej środowiskowej ELEVENLABS_API_KEY",
            status=500,
            mimetype="text/plain"
        )

    data = request.get_json(silent=True) or {}
    text = str(data.get("text", "")).strip()

    if not text:
        return Response(
            "Brak tekstu do przeczytania.",
            status=400,
            mimetype="text/plain"
        )

    voice_id = os.getenv("ELEVENLABS_VOICE_ID", "21m00Tcm4TlvDq8ikWAM")

    url = f"https://api.elevenlabs.io/v1/text-to-speech/{voice_id}"

    payload = {
        "text": text,
        "model_id": "eleven_multilingual_v2",
        "voice_settings": {
            "stability": 0.55,
            "similarity_boost": 0.75,
            "style": 0.2,
            "use_speaker_boost": True
        }
    }

    headers = {
        "xi-api-key": api_key,
        "Content-Type": "application/json",
        "Accept": "audio/mpeg"
    }

    try:
        eleven_response = requests.post(
            url,
            json=payload,
            headers=headers,
            timeout=60
        )

        if not eleven_response.ok:
            return Response(
                eleven_response.text,
                status=eleven_response.status_code,
                mimetype="text/plain"
            )

        return Response(
            eleven_response.content,
            mimetype="audio/mpeg"
        )

    except Exception as error:
        return Response(
            f"Błąd połączenia z ElevenLabs: {error}",
            status=500,
            mimetype="text/plain"
        )


if __name__ == "__main__":
    app.run(debug=True)